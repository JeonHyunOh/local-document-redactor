"""batch_service 단위 테스트 — 폴더 스캔·배치 검색·배치 편집(구조 재현·오류 격리)."""

from __future__ import annotations

from pathlib import Path

import fitz
from openpyxl import Workbook, load_workbook

from document_redactor import batch_service, email_service, file_service
from document_redactor import name_redactor  # noqa: F401  (이름 정리 시나리오에서 사용)
from document_redactor.models import (
    EditRequest,
    ExcelAction,
    SearchCriteria,
    VerificationResult,
)


def _xlsx(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.active["A1"] = value
    wb.save(path)


def _pdf(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), text, fontsize=14)
    doc.save(path)
    doc.close()


def _criteria(*kw: str) -> SearchCriteria:
    return SearchCriteria(keywords=list(kw))


def test_scan_folder_recursive_finds_supported(tmp_path: Path):
    _xlsx(tmp_path / "top.xlsx", "대외비")
    _pdf(tmp_path / "sub" / "deep.pdf", "대외비")
    (tmp_path / "note.txt").write_text("ignored")
    (tmp_path / "~$lock.xlsx").write_bytes(b"lock")  # Office 잠금 파일 제외

    found = batch_service.scan_folder(tmp_path, recursive=True)
    names = {p.name for p in found}
    assert names == {"top.xlsx", "deep.pdf"}


def test_scan_folder_non_recursive_skips_subdirs(tmp_path: Path):
    _xlsx(tmp_path / "top.xlsx", "x")
    _xlsx(tmp_path / "sub" / "nested.xlsx", "x")
    found = batch_service.scan_folder(tmp_path, recursive=False)
    assert {p.name for p in found} == {"top.xlsx"}


def test_batch_search_aggregates_and_isolates_errors(tmp_path: Path):
    _xlsx(tmp_path / "a.xlsx", "대외비 문서")
    _pdf(tmp_path / "b.pdf", "공개 자료")
    (tmp_path / "broken.pdf").write_bytes(b"%PDF-broken not a real pdf")

    items = batch_service.batch_search(tmp_path, _criteria("대외비"), recursive=True)
    by_name = {Path(i.path).name: i for i in items}
    assert by_name["a.xlsx"].matches == 1
    assert by_name["b.pdf"].matches == 0
    assert by_name["broken.pdf"].error is not None  # 오류 격리, 배치는 계속


def test_keyword_summary_and_match_details(tmp_path: Path):
    # PyMuPDF 기본 폰트는 한글 글리프를 못 실으므로 PDF 픽스처는 ASCII 키워드를 쓴다
    # (도구는 PDF를 생성하지 않고 기존 PDF를 검색만 하므로 실제 사용에는 무관).
    _xlsx(tmp_path / "a.xlsx", "대외비 및 내부검토용")  # 한 셀에 두 키워드
    _pdf(tmp_path / "sub" / "b.pdf", "SECRET SECRET report")  # 한 페이지 2회

    items = batch_service.batch_search(tmp_path, _criteria("대외비", "내부검토용", "SECRET"))

    summary = batch_service.keyword_summary(items)
    assert summary["대외비"] == 1
    assert summary["내부검토용"] == 1
    assert summary["SECRET"] == 2  # PDF 한 페이지에서 2회

    details = batch_service.match_details(items)
    # Excel 위치는 '시트!셀', PDF 위치는 'N페이지' 형식
    assert any(r["위치"] == "Sheet!A1" and r["키워드"] == "대외비" for r in details)
    assert any(r["위치"] == "1페이지" and r["키워드"] == "SECRET" for r in details)


def test_batch_search_reports_progress(tmp_path: Path):
    _xlsx(tmp_path / "a.xlsx", "x")
    _xlsx(tmp_path / "b.xlsx", "x")
    calls: list[tuple[int, int, str]] = []
    batch_service.batch_search(
        tmp_path, _criteria("x"), on_progress=lambda d, t, n: calls.append((d, t, n))
    )
    assert [c[0] for c in calls] == [1, 2]  # 완료 카운트 증가
    assert all(c[1] == 2 for c in calls)  # 전체 개수 일정


def test_batch_edit_mirrors_structure_and_verifies(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비 문서")
    _xlsx(root / "sub" / "b.xlsx", "대외비 메모")
    _xlsx(root / "clean.xlsx", "공개 자료")  # 매치 없음 → 사본 생성 안 함
    out = tmp_path / "out"

    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit(root, req, out, recursive=True)
    by_name = {Path(i.path).name: i for i in items}

    # 매치 없는 파일은 출력 없음
    assert by_name["clean.xlsx"].output_path is None

    # 하위 폴더 구조가 재현되어야 함
    edited_sub = out / "sub" / "b_edited.xlsx"
    assert edited_sub.exists()
    assert load_workbook(edited_sub).active["A1"].value is None

    # 편집된 파일은 모두 재검증 clean
    edited = [i for i in items if i.output_path]
    assert edited and all(i.clean for i in edited)

    # 원본은 그대로
    assert load_workbook(root / "a.xlsx").active["A1"].value == "대외비 문서"


def test_batch_edit_in_place_replaces_and_backs_up(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비 문서")
    _xlsx(root / "sub" / "b.xlsx", "대외비 메모")
    _xlsx(root / "clean.xlsx", "공개 자료")  # 매치 없음 → 교체·백업 안 함
    backup = tmp_path / "src_backup"

    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)
    by_name = {Path(i.path).name: i for i in items}

    # 원본이 제자리에서 편집됨
    assert load_workbook(root / "a.xlsx").active["A1"].value is None
    assert load_workbook(root / "sub" / "b.xlsx").active["A1"].value is None
    # 파일 경로/이름은 그대로 (접미사 없음)
    assert by_name["a.xlsx"].output_path == str(root / "a.xlsx")

    # 백업에 원본 내용이 구조 그대로 보존됨
    assert load_workbook(backup / "a.xlsx").active["A1"].value == "대외비 문서"
    assert load_workbook(backup / "sub" / "b.xlsx").active["A1"].value == "대외비 메모"

    # 매치 없는 파일은 교체·백업 안 함
    assert by_name["clean.xlsx"].output_path is None
    assert not (backup / "clean.xlsx").exists()


def test_batch_edit_in_place_keeps_original_when_verify_fails(tmp_path: Path, monkeypatch):
    """재검증 실패 시 원본을 덮어쓰지 않고 백업도 만들지 않는다(안전)."""
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비 문서")
    backup = tmp_path / "src_backup"

    # 재검증이 항상 실패한다고 가정
    monkeypatch.setattr(
        file_service,
        "verify",
        lambda out, crit: VerificationResult(output_path=str(out), clean=False),
    )

    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    # 원본 그대로, 교체 안 됨
    assert load_workbook(root / "a.xlsx").active["A1"].value == "대외비 문서"
    # 백업도 만들지 않음
    assert not backup.exists()
    # 결과에 실패 사유 기록
    assert items[0].error and "재검증 실패" in items[0].error


# --------------------------------------------------------------------------- #
# 파일명·폴더명 키워드 정리 — batch_edit(출력본)
# --------------------------------------------------------------------------- #
def test_batch_edit_redacts_output_names(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "포스코_a.xlsx", "대외비 문서")                 # 내용+이름 매치
    _xlsx(root / "포스코_폴더" / "b.xlsx", "대외비 메모")        # 폴더명 매치 + 내용 매치
    out = tmp_path / "out"

    req = EditRequest(criteria=_criteria("대외비", "포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit(root, req, out, recursive=True)

    # 최상위 파일명에서 '포스코' 제거 (편집본 접미사는 유지)
    assert (out / "a_edited.xlsx").exists()
    # 폴더명에서 '포스코' 제거되어 정리된 폴더에 저장
    assert (out / "폴더" / "b_edited.xlsx").exists()
    # 원본은 그대로
    assert (root / "포스코_a.xlsx").exists()


def test_batch_edit_copies_name_only_match(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "포스코_보고서.xlsx", "공개 자료")  # 내용 깨끗, 파일명만 매치
    out = tmp_path / "out"

    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit(root, req, out, recursive=True)
    item = items[0]

    # 정리된 이름으로 복사됨(내용 무수정)
    copied = out / "보고서.xlsx"
    assert copied.exists()
    assert load_workbook(copied).active["A1"].value == "공개 자료"
    assert item.renamed_to == "보고서.xlsx"


def test_batch_edit_skips_when_content_and_name_clean(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "일반.xlsx", "공개 자료")
    out = tmp_path / "out"

    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit(root, req, out, recursive=True)

    assert items[0].output_path is None
    assert not out.exists() or not any(out.rglob("*.xlsx"))


def test_batch_edit_name_collision_gets_suffix(tmp_path: Path):
    root = tmp_path / "src"
    # 서로 다른 두 파일이 정리 후 같은 이름('보고서.xlsx')이 됨 — 둘 다 이름만 매치(내용 깨끗)
    _xlsx(root / "포스코_보고서.xlsx", "공개 자료")
    _xlsx(root / "보고서_포스코.xlsx", "공개 자료")
    out = tmp_path / "out"

    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit(root, req, out, recursive=True)

    names = {p.name for p in out.glob("*.xlsx")}
    assert names == {"보고서.xlsx", "보고서_1.xlsx"}


# --------------------------------------------------------------------------- #
# 파일명·폴더명 키워드 정리 — batch_edit_in_place(제자리 rename)
# --------------------------------------------------------------------------- #
def test_batch_edit_in_place_renames_name_only_file(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "포스코_보고서.xlsx", "공개 자료")  # 내용 깨끗, 파일명만 매치
    backup = tmp_path / "src_backup"

    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    # 파일이 정리된 이름으로 rename됨, 원래 이름은 사라짐
    assert (root / "보고서.xlsx").exists()
    assert not (root / "포스코_보고서.xlsx").exists()
    # 내용은 무수정
    assert load_workbook(root / "보고서.xlsx").active["A1"].value == "공개 자료"
    # rename 전 백업 생성
    assert (backup / "포스코_보고서.xlsx").exists()
    # rename 로그 기록
    log = (backup / "_rename_log.txt").read_text(encoding="utf-8")
    assert "포스코_보고서.xlsx" in log and "보고서.xlsx" in log
    assert items[0].renamed_to == "보고서.xlsx"


def test_batch_edit_in_place_renames_folder_bottom_up_not_root(tmp_path: Path):
    root = tmp_path / "포스코_루트"          # 루트 이름에도 키워드 → 변경되면 안 됨
    _xlsx(root / "포스코_하위" / "a.xlsx", "대외비 문서")
    backup = tmp_path / "포스코_루트_backup"

    req = EditRequest(criteria=_criteria("대외비", "포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    # 하위 폴더는 rename, 루트는 그대로
    assert root.exists()                      # 루트 미변경
    assert (root / "하위").is_dir()            # 하위 폴더 '포스코' 제거
    assert not (root / "포스코_하위").exists()
    assert (root / "하위" / "a.xlsx").exists()  # 파일도 함께 이동


def test_batch_edit_in_place_content_and_name_match(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "포스코_a.xlsx", "대외비 문서")  # 내용+이름 둘 다 매치
    backup = tmp_path / "src_backup"

    req = EditRequest(criteria=_criteria("대외비", "포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    # 내용 편집 + 파일명 정리 모두 적용
    assert (root / "a.xlsx").exists()
    assert load_workbook(root / "a.xlsx").active["A1"].value is None
    # 백업엔 원본 이름·원본 내용 보존
    assert load_workbook(backup / "포스코_a.xlsx").active["A1"].value == "대외비 문서"


# --------------------------------------------------------------------------- #
# 이메일(.msg/.eml) 배치 처리
# --------------------------------------------------------------------------- #
def test_scan_folder_includes_email(tmp_path: Path, make_eml):
    _xlsx(tmp_path / "a.xlsx", "x")
    make_eml(tmp_path / "b.eml", subject="s", body="x")
    names = {p.name for p in batch_service.scan_folder(tmp_path, recursive=True)}
    assert names == {"a.xlsx", "b.eml"}


def test_batch_edit_email_produces_markdown(tmp_path: Path, make_eml):
    root = tmp_path / "src"
    make_eml(root / "메일.eml", subject="포스코 보고", body="대외비 내용")
    out = tmp_path / "out"
    req = EditRequest(criteria=_criteria("포스코", "대외비"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit(root, req, out, recursive=True)

    produced = out / "메일_redacted.md"
    assert produced.exists()
    text = produced.read_text(encoding="utf-8")
    assert "포스코" not in text and "대외비" not in text
    assert any(i.output_path for i in items)
    # 원본 보존
    assert (root / "메일.eml").exists()


def test_batch_edit_in_place_email_produces_md_and_deletes_original(tmp_path: Path, make_eml):
    root = tmp_path / "src"
    make_eml(root / "포스코_메일.eml", subject="포스코", body="대외비")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("포스코", "대외비"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)
    item = items[0]

    # 원본 .eml은 삭제되고, 정리된 .md가 생성됨(파일명 키워드도 정리)
    assert not (root / "포스코_메일.eml").exists()
    md = root / "메일_redacted.md"
    assert md.exists()
    text = md.read_text(encoding="utf-8")
    assert "포스코" not in text and "대외비" not in text
    # 원본은 백업에 보존(복구 가능)
    assert (backup / "포스코_메일.eml").exists()
    assert item.output_path == str(md)
    assert item.renamed_to == "메일_redacted.md"


def test_batch_edit_in_place_clean_email_untouched(tmp_path: Path, make_eml):
    # 키워드가 전혀 없는 이메일은 변환·삭제하지 않는다.
    root = tmp_path / "src"
    make_eml(root / "공지.eml", subject="공지", body="일반 안내")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert (root / "공지.eml").exists()  # 그대로 유지
    assert not any(p.suffix == ".md" for p in root.rglob("*"))


# --------------------------------------------------------------------------- #
# 형식 제거(.dwg/.png/.nwd) + 확장자 무관 파일명 정리 + pptx 배치
# --------------------------------------------------------------------------- #
def test_scan_all_files_returns_every_extension(tmp_path: Path):
    _xlsx(tmp_path / "a.xlsx", "x")
    (tmp_path / "b.dwg").write_bytes(b"dwg")
    (tmp_path / "c.txt").write_text("t", encoding="utf-8")
    (tmp_path / "~$lock.xlsx").write_bytes(b"lock")
    names = {p.name for p in batch_service.scan_all_files(tmp_path, recursive=True)}
    assert names == {"a.xlsx", "b.dwg", "c.txt"}


def test_in_place_removes_target_suffixes_hard_delete(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비 문서")
    (root / "도면.dwg").write_bytes(b"dwg")
    (root / "이미지.png").write_bytes(b"png")
    (root / "모델.nwd").write_bytes(b"nwd")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)

    items = batch_service.batch_edit_in_place(
        root, req, backup, recursive=True, remove_suffixes={".dwg", ".png", ".nwd"}
    )
    # 완전 삭제됨(백업 없음)
    assert not (root / "도면.dwg").exists()
    assert not (root / "이미지.png").exists()
    assert not (root / "모델.nwd").exists()
    assert not (backup / "도면.dwg").exists()
    # 삭제 로그는 대상 폴더 안에 저장하고, 확장자별 개수만 기록(파일명 미기록)
    assert not (backup / "_removed_log.txt").exists()  # 백업 폴더가 아니라
    log = (root / "_removed_log.txt").read_text(encoding="utf-8")  # 대상 폴더에 저장
    assert "도면" not in log and "이미지" not in log and "모델" not in log  # 파일명 미기록
    assert ".dwg 1개" in log and ".png 1개" in log and ".nwd 1개" in log     # 확장자별 개수
    # 결과에 note
    assert any(i.note and "완전 삭제" in i.note for i in items)
    # 지원 파일은 정상 편집
    assert load_workbook(root / "a.xlsx").active["A1"].value is None


def test_removed_log_not_mangled_when_keyword_matches_its_name(tmp_path: Path):
    # 키워드가 '_removed_log.txt' 이름과 겹쳐도 로그 파일은 Phase 2a rename 대상이 아니다.
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비")
    (root / "도면.dwg").write_bytes(b"dwg")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("대외비", "log"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True, remove_suffixes={".dwg"})

    assert (root / "_removed_log.txt").exists()  # 원래 이름 그대로 유지
    assert not (root / "removed_log.txt").exists()  # 훼손된 이름 없음


def test_in_place_no_spurious_rename_when_keyword_only_in_extension(tmp_path: Path):
    # 키워드가 확장자에만 있으면(정리해도 이름 불변) rename하지 않는다(자기 충돌 방지).
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("대외비", "xlsx"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert (root / "a.xlsx").exists()          # 그대로
    assert not (root / "a_1.xlsx").exists()      # 헛된 접미사 rename 없음
    assert all(i.renamed_to is None for i in items)


def test_in_place_removal_off_by_default(tmp_path: Path):
    root = tmp_path / "src"
    _xlsx(root / "a.xlsx", "대외비")
    (root / "도면.dwg").write_bytes(b"dwg")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)  # remove_suffixes 없음
    assert (root / "도면.dwg").exists()  # 삭제 안 됨


def test_in_place_renames_unsupported_extension_by_filename(tmp_path: Path):
    root = tmp_path / "src"
    root.mkdir(parents=True)
    (root / "포스코_메모.txt").write_text("공개", encoding="utf-8")  # 미지원 형식, 파일명만 키워드
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert (root / "메모.txt").exists()  # 확장자 무관 파일명 정리
    assert not (root / "포스코_메모.txt").exists()
    assert any(i.renamed_to == "메모.txt" for i in items)


def _make_zip(path: Path, inner_name: str, data: bytes) -> None:
    import zipfile

    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(inner_name, data)


def test_in_place_extracts_zip_and_deletes_original(tmp_path: Path):
    import io

    from openpyxl import Workbook, load_workbook

    root = tmp_path / "src"
    root.mkdir(parents=True)
    buf = io.BytesIO()
    wb = Workbook(); wb.active["A1"] = "대외비"; wb.save(buf)
    _make_zip(root / "압축.zip", "포스코_내부.xlsx", buf.getvalue())
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("대외비", "포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert not (root / "압축.zip").exists()      # 원본 zip 삭제
    assert (backup / "압축.zip").exists()          # 백업 보존
    extracted = root / "압축" / "내부.xlsx"        # 파일명 '포스코' 정리됨
    assert extracted.exists()
    assert load_workbook(extracted).active["A1"].value is None  # 내용 편집됨


def test_in_place_zip_slip_rejected(tmp_path: Path):
    root = tmp_path / "src"
    root.mkdir(parents=True)
    _make_zip(root / "evil.zip", "../escape.txt", b"x")
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("x"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert not (tmp_path / "escape.txt").exists()  # 상위로 탈출 안 됨
    assert any(i.error for i in items if i.path.endswith("evil.zip"))


def test_in_place_pptx_content_edited(tmp_path: Path, make_pptx):
    root = tmp_path / "src"
    make_pptx(root / "deck.pptx", title="포스코 제목", body_lines=["대외비 본문"])
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("포스코", "대외비"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    from document_redactor import pptx_service
    assert pptx_service.search(root / "deck.pptx", req.criteria).total_matches == 0
    assert (backup / "deck.pptx").exists()  # 원본 백업


def _fake_office_pdf(tmp: Path, text: str):
    import fitz
    p = tmp / "conv_office.pdf"
    doc = fitz.open(); pg = doc.new_page(); pg.insert_text((72, 72), text); doc.save(str(p)); doc.close()
    return p


def test_in_place_office_doc_converts_to_pdf_and_deletes_original(tmp_path: Path, monkeypatch):
    from document_redactor import doc_converter
    root = tmp_path / "src"
    root.mkdir(parents=True)
    (root / "포스코_보고서.docx").write_bytes(b"stub")
    monkeypatch.setattr(doc_converter, "convert_to_pdf",
                        lambda p: _fake_office_pdf(tmp_path, "SECRET 010-1234-5678"))
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("SECRET", "포스코"), excel_action=ExcelAction.CLEAR_CELL)
    items = batch_service.batch_edit_in_place(root, req, backup, recursive=True)
    item = items[0]

    assert not (root / "포스코_보고서.docx").exists()          # 원본 삭제
    assert (root / "보고서_redacted.pdf").exists()             # 변환·정리 PDF(파일명도 정리)
    assert (backup / "포스코_보고서.docx").exists()            # 원본 백업
    assert item.renamed_to == "보고서_redacted.pdf"
    # 산출 PDF 재검증 clean
    import fitz
    assert "SECRET" not in fitz.open(str(root / "보고서_redacted.pdf"))[0].get_text()


def test_in_place_clean_office_doc_untouched(tmp_path: Path, monkeypatch):
    from document_redactor import doc_converter
    root = tmp_path / "src"
    root.mkdir(parents=True)
    (root / "공지.docx").write_bytes(b"stub")
    monkeypatch.setattr(doc_converter, "convert_to_pdf",
                        lambda p: _fake_office_pdf(tmp_path, "general notice"))
    backup = tmp_path / "src_backup"
    req = EditRequest(criteria=_criteria("포스코"), excel_action=ExcelAction.CLEAR_CELL)
    batch_service.batch_edit_in_place(root, req, backup, recursive=True)

    assert (root / "공지.docx").exists()  # 키워드·패턴 없음 → 유지
    assert not any(p.suffix == ".pdf" for p in root.rglob("*"))
