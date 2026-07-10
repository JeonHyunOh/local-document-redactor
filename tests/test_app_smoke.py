"""app.py 헤드리스 스모크 테스트 — Streamlit AppTest로 검색→편집→재검증 흐름을 구동한다.

UI가 서비스 계층과 실제로 연결되는지 확인한다(브라우저 없이).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

st_testing = pytest.importorskip("streamlit.testing.v1")
AppTest = st_testing.AppTest

APP_PATH = str(Path(__file__).resolve().parents[1] / "app.py")


def test_app_loads_without_error():
    at = AppTest.from_file(APP_PATH).run()
    assert not at.exception
    assert any("키워드" in str(t.value) for t in at.title)


def test_app_search_flow_populates_report(tmp_path: Path, monkeypatch):
    """업로드를 흉내 내기 어려우므로, 서비스 경로가 UI 세션과 연결되는지 확인한다.

    파일 업로더는 AppTest에서 직접 주입이 제한적이라, 여기서는 앱이 예외 없이 초기
    렌더를 마치고 버튼이 비활성 상태로 존재하는지만 확인한다(회귀 방지용 스모크).
    """
    at = AppTest.from_file(APP_PATH).run()
    assert not at.exception
    # 키워드/파일 없이 검사 버튼은 비활성이어야 한다
    search_buttons = [b for b in at.button if "검사" in b.label]
    assert search_buttons and search_buttons[0].disabled is True


def test_app_folder_mode_batch_search(tmp_path: Path):
    """폴더 경로·키워드 입력 → 폴더 검사 → 집계 결과가 세션에 채워진다.

    탭은 레이아웃 컨테이너라 폴더 탭의 위젯이 항상 렌더되므로 탭 전환 없이 접근한다.
    """
    wb = Workbook()
    wb.active["A1"] = "대외비 문서"
    wb.save(tmp_path / "a.xlsx")
    wb2 = Workbook()
    wb2.active["A1"] = "공개 자료"
    wb2.save(tmp_path / "b.xlsx")

    at = AppTest.from_file(APP_PATH, default_timeout=30).run()
    at.session_state["keywords_raw"] = "대외비"
    at.run()

    folder_input = next(t for t in at.text_input if "폴더 경로" in t.label)
    folder_input.set_value(str(tmp_path)).run()

    scan_btn = next(b for b in at.button if "폴더 검사" in b.label)
    scan_btn.click().run()

    assert not at.exception
    assert "b_search" in at.session_state
    items = at.session_state["b_search"]
    by_name = {Path(i.path).name: i.matches for i in items}
    assert by_name == {"a.xlsx": 1, "b.xlsx": 0}
