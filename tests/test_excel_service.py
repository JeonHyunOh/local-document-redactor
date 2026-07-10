"""excel_service 단위 테스트 (M2). 픽스처는 런타임에 openpyxl로 생성한다."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from document_redactor import excel_service
from document_redactor.models import (
    EditRequest,
    ExcelAction,
    SearchCriteria,
    SearchMode,
)


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """두 시트에 키워드가 흩어진 샘플 워크북을 만든다."""
    wb = Workbook()
    s1 = wb.active
    s1.title = "Sheet1"
    s1["A1"] = "대외비 문서"
    s1["B1"] = "공개 자료"
    s1["A2"] = "내부검토용 메모"
    s1["C3"] = "=SUM(1,2)"  # 수식 셀 — 검사 제외 대상
    s2 = wb.create_sheet("Sheet2")
    s2["A1"] = "대외비 및 내부검토용"  # 한 행에 두 키워드
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


def _criteria(*keywords: str, mode: SearchMode = SearchMode.CONTAINS) -> SearchCriteria:
    return SearchCriteria(keywords=list(keywords), mode=mode, case_sensitive=False)


# --------------------------------------------------------------------------- #
# 검색
# --------------------------------------------------------------------------- #
def test_search_across_multiple_sheets(sample_xlsx: Path):
    report = excel_service.search(sample_xlsx, _criteria("대외비", "내부검토용"))
    sheets = {m.sheet_name for m in report.excel_matches}
    assert sheets == {"Sheet1", "Sheet2"}
    # 대외비: Sheet1!A1, Sheet2!A1 / 내부검토용: Sheet1!A2, Sheet2!A1 → 4건
    assert len(report.excel_matches) == 4


def test_search_skips_formula_cells(sample_xlsx: Path):
    report = excel_service.search(sample_xlsx, _criteria("SUM"))
    assert report.excel_matches == []


def test_search_no_keyword_returns_empty(sample_xlsx: Path):
    report = excel_service.search(sample_xlsx, _criteria("없는키워드"))
    assert report.total_matches == 0


# --------------------------------------------------------------------------- #
# 편집: 키워드만 제거
# --------------------------------------------------------------------------- #
def test_remove_keyword_only(sample_xlsx: Path, tmp_path: Path):
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.REMOVE_KEYWORD)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out")
    wb = load_workbook(result.output_path)
    assert wb["Sheet1"]["A1"].value == " 문서"
    assert wb["Sheet2"]["A1"].value == " 및 내부검토용"


# --------------------------------------------------------------------------- #
# 편집: 셀 전체 비우기
# --------------------------------------------------------------------------- #
def test_clear_cell(sample_xlsx: Path, tmp_path: Path):
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out")
    wb = load_workbook(result.output_path)
    assert wb["Sheet1"]["A1"].value is None
    assert wb["Sheet2"]["A1"].value is None
    assert result.cells_changed == 2


# --------------------------------------------------------------------------- #
# 편집: 행 전체 삭제 + 같은 행 중복 방지
# --------------------------------------------------------------------------- #
def test_delete_row_dedupes_same_row(sample_xlsx: Path, tmp_path: Path):
    # Sheet2!A1은 두 키워드를 모두 포함 → 행 1건만 삭제되어야 한다.
    req = EditRequest(
        criteria=_criteria("대외비", "내부검토용"), excel_action=ExcelAction.DELETE_ROW
    )
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out")
    wb = load_workbook(result.output_path)
    # Sheet1: 행1(대외비), 행2(내부검토용) 삭제 → C3 수식만 남아 위로 이동
    # Sheet2: 행1 삭제 → 비어야 함
    assert wb["Sheet2"]["A1"].value is None
    assert result.rows_deleted == 3  # Sheet1 두 행 + Sheet2 한 행


def test_delete_row_descending_preserves_other_rows(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "keep-top"
    ws["A2"] = "삭제대상 행"
    ws["A3"] = "삭제대상 다시"
    ws["A4"] = "keep-bottom"
    path = tmp_path / "rows.xlsx"
    wb.save(path)

    req = EditRequest(criteria=_criteria("삭제대상"), excel_action=ExcelAction.DELETE_ROW)
    result = excel_service.apply_edit(path, req, tmp_path / "out")
    out = load_workbook(result.output_path).active
    assert [out.cell(r, 1).value for r in range(1, 3)] == ["keep-top", "keep-bottom"]
    assert result.rows_deleted == 2


# --------------------------------------------------------------------------- #
# 원본 보존
# --------------------------------------------------------------------------- #
def test_original_file_unchanged(sample_xlsx: Path, tmp_path: Path):
    before = sample_xlsx.read_bytes()
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out")
    assert sample_xlsx.read_bytes() == before  # 원본 바이트 동일
    assert Path(result.output_path) != sample_xlsx
    assert Path(result.output_path).name == "sample_edited.xlsx"


# --------------------------------------------------------------------------- #
# 재검증
# --------------------------------------------------------------------------- #
def test_verify_after_edit_is_clean(sample_xlsx: Path, tmp_path: Path):
    criteria = _criteria("대외비", "내부검토용")
    req = EditRequest(criteria=criteria, excel_action=ExcelAction.CLEAR_CELL)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out")
    verification = excel_service.verify(Path(result.output_path), criteria)
    assert verification.clean is True
    assert verification.remaining is None


# --------------------------------------------------------------------------- #
# 미리보기 + 미지원 형식
# --------------------------------------------------------------------------- #
def test_preview_fills_expected_value(sample_xlsx: Path):
    report = excel_service.search(sample_xlsx, _criteria("대외비"))
    previewed = excel_service.preview(report, ExcelAction.REMOVE_KEYWORD)
    a1 = next(m for m in previewed if m.cell == "A1" and m.sheet_name == "Sheet1")
    assert a1.expected_value == " 문서"


def test_unsupported_extension_rejected(tmp_path: Path):
    bad = tmp_path / "legacy.xls"
    bad.write_bytes(b"not really xls")
    with pytest.raises(ValueError, match="지원하지 않는"):
        excel_service.search(bad, _criteria("x"))


# --------------------------------------------------------------------------- #
# 선택 삭제 (selected)
# --------------------------------------------------------------------------- #
def test_selected_remove_keyword_only_selected(tmp_path: Path):
    wb = Workbook()
    wb.active["A1"] = "대외비 및 내부검토용"  # 한 셀에 두 키워드
    path = tmp_path / "s.xlsx"
    wb.save(path)

    report = excel_service.search(path, _criteria("대외비", "내부검토용"))
    chosen = [m for m in report.excel_matches if m.keyword == "대외비"]  # 대외비만 선택
    req = EditRequest(criteria=_criteria("대외비", "내부검토용"), excel_action=ExcelAction.REMOVE_KEYWORD)
    result = excel_service.apply_edit(path, req, tmp_path / "out", selected=chosen)

    # 선택한 '대외비'만 제거되고 '내부검토용'은 남아야 한다
    assert load_workbook(result.output_path).active["A1"].value == " 및 내부검토용"


def test_selected_clear_cell_only_selected(sample_xlsx: Path, tmp_path: Path):
    report = excel_service.search(sample_xlsx, _criteria("대외비"))
    chosen = [m for m in report.excel_matches if m.sheet_name == "Sheet1"]  # Sheet1만 선택
    req = EditRequest(criteria=_criteria("대외비"), excel_action=ExcelAction.CLEAR_CELL)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out", selected=chosen)

    wb = load_workbook(result.output_path)
    assert wb["Sheet1"]["A1"].value is None            # 선택됨 → 비워짐
    assert wb["Sheet2"]["A1"].value == "대외비 및 내부검토용"  # 미선택 → 유지
    assert result.cells_changed == 1


def test_selected_delete_row_only_selected(sample_xlsx: Path, tmp_path: Path):
    report = excel_service.search(sample_xlsx, _criteria("대외비", "내부검토용"))
    chosen = [m for m in report.excel_matches if m.sheet_name == "Sheet2"]  # Sheet2 행만
    req = EditRequest(criteria=_criteria("대외비", "내부검토용"), excel_action=ExcelAction.DELETE_ROW)
    result = excel_service.apply_edit(sample_xlsx, req, tmp_path / "out", selected=chosen)

    wb = load_workbook(result.output_path)
    assert wb["Sheet2"]["A1"].value is None                 # 선택 행 삭제
    assert wb["Sheet1"]["A1"].value == "대외비 문서"          # 미선택 시트 유지
    assert result.rows_deleted == 1
