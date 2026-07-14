"""Excel(.xlsx/.xlsm) 검색·편집 서비스.

검색(파일 무수정)과 편집(승인 후)을 분리한다. 규칙:
- 모든 워크시트 셀 검사. **수식 셀의 수식 문자열은 검사·편집 대상에서 제외**(계산 결과 손상 방지).
- 문자열 값 셀만 검사한다(숫자·날짜는 키워드 텍스트 검색 대상에서 제외).
- DELETE_ROW은 같은 행 중복 삭제 방지 + **아래쪽 행부터(내림차순) 삭제**.
- .xlsm은 keep_vba=True로 로드. 원본을 덮어쓰지 않고 ``_edited`` 접미사 새 파일 생성.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from . import keyword_matcher, pattern_matcher
from .models import (
    EditRequest,
    EditResult,
    ExcelAction,
    ExcelMatch,
    FileType,
    SearchCriteria,
    SearchReport,
    VerificationResult,
)

_SUPPORTED_SUFFIXES = {".xlsx": FileType.XLSX, ".xlsm": FileType.XLSM}


def _resolve_file_type(path: Path) -> FileType:
    """확장자로 지원 유형을 판정한다. 미지원이면 명확히 거부한다."""
    suffix = path.suffix.lower()
    if suffix not in _SUPPORTED_SUFFIXES:
        raise ValueError(
            f"지원하지 않는 Excel 형식입니다: {suffix or '(확장자 없음)'}. "
            "지원 형식은 .xlsx, .xlsm 입니다."
        )
    return _SUPPORTED_SUFFIXES[suffix]


def _searchable_text(cell: Cell) -> str | None:
    """검사 대상 문자열을 반환한다. 수식 셀·비문자열 값은 None(제외)."""
    if cell.data_type == "f":  # 수식 셀은 제외
        return None
    if isinstance(cell.value, str):
        return cell.value
    return None


def search(path: Path, criteria: SearchCriteria) -> SearchReport:
    """Excel 파일을 수정하지 않고 키워드를 검사해 결과를 반환한다.

    read_only로 열어 원본을 절대 건드리지 않는다. expected_value는 여기서 채우지
    않으며(삭제 방식은 검색 이후 선택), 미리보기는 :func:`preview` 참조.
    """
    file_type = _resolve_file_type(path)
    keep_vba = file_type is FileType.XLSM
    workbook = load_workbook(path, read_only=True, keep_vba=keep_vba, data_only=False)
    try:
        matches: list[ExcelMatch] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    text = _searchable_text(cell)
                    if text is None:
                        continue
                    for keyword in keyword_matcher.find_matches(text, criteria):
                        matches.append(
                            ExcelMatch(
                                file_name=path.name,
                                sheet_name=sheet.title,
                                cell=cell.coordinate,
                                row=cell.row,
                                keyword=keyword,
                                original_value=text,
                            )
                        )
                    for label, _value in pattern_matcher.find_patterns(
                        text, criteria.redact_account_numbers
                    ):
                        matches.append(
                            ExcelMatch(
                                file_name=path.name,
                                sheet_name=sheet.title,
                                cell=cell.coordinate,
                                row=cell.row,
                                keyword=f"[{label}]",  # 패턴은 유형 라벨로 구분
                                original_value=text,
                            )
                        )
    finally:
        workbook.close()

    return SearchReport(
        file_name=path.name,
        file_type=file_type,
        criteria=criteria,
        excel_matches=matches,
    )


def preview(report: SearchReport, action: ExcelAction) -> list[ExcelMatch]:
    """검색 결과에 선택된 삭제 방식의 예상 변경 값을 채워 반환한다(UI 미리보기용)."""
    previewed: list[ExcelMatch] = []
    for match in report.excel_matches:
        if action is ExcelAction.REMOVE_KEYWORD:
            expected = keyword_matcher.remove_keywords(
                match.original_value,
                report.criteria.keywords,
                report.criteria.case_sensitive,
            )
        elif action is ExcelAction.CLEAR_CELL:
            expected = ""
        else:  # DELETE_ROW
            expected = "(행 삭제)"
        previewed.append(match.model_copy(update={"expected_value": expected}))
    return previewed


def apply_edit(
    path: Path,
    request: EditRequest,
    output_dir: Path,
    selected: list[ExcelMatch] | None = None,
) -> EditResult:
    """승인된 편집을 실행해 ``_edited`` 접미사가 붙은 새 파일을 생성한다.

    원본을 덮어쓰지 않는다. DELETE_ROW은 시트별로 행을 중복 없이 모아 내림차순
    삭제해 행 번호 밀림을 방지한다.

    selected가 주어지면 **그 항목(시트·셀·키워드)만** 처리한다. None이면 검색 조건에
    매칭되는 모든 항목을 처리한다.
    """
    file_type = _resolve_file_type(path)
    keep_vba = file_type is FileType.XLSM
    criteria = request.criteria
    action = request.excel_action

    workbook = load_workbook(path, keep_vba=keep_vba, data_only=False)
    log: list[str] = []
    cells_changed = 0
    rows_deleted = 0

    try:
        if selected is not None:
            cells_changed, rows_deleted = _apply_selected(workbook, action, criteria, selected, log)
        else:
            cells_changed, rows_deleted = _apply_all(workbook, action, criteria, log)

        # 패턴은 키워드 선택·삭제 방식과 무관하게 항상 부분 제거한다(개인정보 자동 삭제).
        cells_changed += _apply_patterns_all(workbook, criteria, log)

        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{path.stem}_edited{path.suffix}"
        workbook.save(output_path)
    finally:
        workbook.close()

    return EditResult(
        source_name=path.name,
        output_path=str(output_path),
        file_type=file_type,
        cells_changed=cells_changed,
        rows_deleted=rows_deleted,
        log=log,
    )


def _apply_patterns_all(workbook, criteria, log: list[str]) -> int:
    """모든 문자열 셀에서 정형 개인정보 패턴을 부분 제거한다(항상 적용). 변경 셀 수 반환."""
    include_account = criteria.redact_account_numbers
    changed = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                text = _searchable_text(cell)
                if text is None:
                    continue
                new_value = pattern_matcher.remove_patterns(text, include_account)
                if new_value != text:
                    cell.value = new_value
                    changed += 1
                    log.append(f"{sheet.title}!{cell.coordinate}: 패턴 제거")
    return changed


def _apply_all(workbook, action, criteria, log: list[str]) -> tuple[int, int]:
    """검색 조건에 매칭되는 모든 셀을 처리한다(기존 동작)."""
    cells_changed = 0
    rows_deleted = 0
    for sheet in workbook.worksheets:
        rows_to_delete: set[int] = set()
        for row in sheet.iter_rows():
            for cell in row:
                text = _searchable_text(cell)
                if text is None:
                    continue
                if not keyword_matcher.find_matches(text, criteria):
                    continue

                if action is ExcelAction.REMOVE_KEYWORD:
                    new_value = keyword_matcher.remove_keywords(
                        text, criteria.keywords, criteria.case_sensitive
                    )
                    if new_value != text:
                        cell.value = new_value
                        cells_changed += 1
                        log.append(f"{sheet.title}!{cell.coordinate}: 키워드 제거")
                elif action is ExcelAction.CLEAR_CELL:
                    cell.value = None
                    cells_changed += 1
                    log.append(f"{sheet.title}!{cell.coordinate}: 셀 비움")
                else:  # DELETE_ROW — 같은 행 중복 방지
                    rows_to_delete.add(cell.row)

        if action is ExcelAction.DELETE_ROW and rows_to_delete:
            for row_idx in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_idx, 1)
                rows_deleted += 1
                log.append(f"{sheet.title}!행{row_idx}: 행 삭제")
    return cells_changed, rows_deleted


def _apply_selected(
    workbook, action, criteria, selected: list[ExcelMatch], log: list[str]
) -> tuple[int, int]:
    """선택된 항목(시트·셀·키워드)만 처리한다.

    - REMOVE_KEYWORD: 같은 셀에서 선택된 키워드만 제거.
    - CLEAR_CELL: 선택된 셀만 비움.
    - DELETE_ROW: 선택된 항목이 속한 행만 중복 없이 내림차순 삭제.
    """
    cells_changed = 0
    rows_deleted = 0

    if action is ExcelAction.REMOVE_KEYWORD:
        by_cell: dict[tuple[str, str], set[str]] = {}
        for m in selected:
            by_cell.setdefault((m.sheet_name, m.cell), set()).add(m.keyword)
        for (sheet_name, coord), keywords in by_cell.items():
            cell = workbook[sheet_name][coord]
            if not isinstance(cell.value, str):
                continue
            new_value = keyword_matcher.remove_keywords(
                cell.value, list(keywords), criteria.case_sensitive
            )
            if new_value != cell.value:
                cell.value = new_value
                cells_changed += 1
                log.append(f"{sheet_name}!{coord}: 선택 키워드 제거")

    elif action is ExcelAction.CLEAR_CELL:
        seen: set[tuple[str, str]] = set()
        for m in selected:
            key = (m.sheet_name, m.cell)
            if key in seen:
                continue
            seen.add(key)
            workbook[m.sheet_name][m.cell].value = None
            cells_changed += 1
            log.append(f"{m.sheet_name}!{m.cell}: 셀 비움")

    else:  # DELETE_ROW
        rows_by_sheet: dict[str, set[int]] = {}
        for m in selected:
            rows_by_sheet.setdefault(m.sheet_name, set()).add(m.row)
        for sheet_name, rows in rows_by_sheet.items():
            sheet = workbook[sheet_name]
            for row_idx in sorted(rows, reverse=True):
                sheet.delete_rows(row_idx, 1)
                rows_deleted += 1
                log.append(f"{sheet_name}!행{row_idx}: 행 삭제")

    return cells_changed, rows_deleted


def verify(output_path: Path, criteria: SearchCriteria) -> VerificationResult:
    """편집된 파일을 재검색해 키워드 잔존 여부를 검증한다."""
    report = search(output_path, criteria)
    return VerificationResult(
        output_path=str(output_path),
        clean=report.total_matches == 0,
        remaining=report if report.total_matches else None,
    )
